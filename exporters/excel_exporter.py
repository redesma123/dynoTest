"""
Excel Exporter untuk DynoTest & BrakeTest.
Menghasilkan file .xlsx profesional menggunakan openpyxl.
"""

from datetime import datetime
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import BrakeResult, DynoResult, EvaluationStatus, TestSession, Vehicle

COLOR_HEADER_BG = "0F172A"       # Dark Navy Slate
COLOR_HEADER_TEXT = "F8FAFC"     # White
COLOR_SUBHEADER_BG = "F1F5F9"    # Light Slate
COLOR_ACCENT = "0284C7"          # Sky Blue
COLOR_PASS_BG = "DCFCE7"         # Light Green
COLOR_PASS_TEXT = "166534"
COLOR_FAIL_BG = "FEE2E2"         # Light Red
COLOR_FAIL_TEXT = "991B1B"
COLOR_BORDER = "CBD5E1"


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def export_dyno_to_excel(
    filepath: str,
    session: Optional[TestSession],
    vehicle: Optional[Vehicle],
    result: DynoResult,
) -> bool:
    """Export hasil pengujian Dyno Test ke file Excel."""
    wb = openpyxl.Workbook()
    
    # ── SHEET 1: RINGKASAN HASIL ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ringkasan Dyno Test"
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1["A1"] = "AUTO-TECH SYSTEMS — LAPORAN PENGUJIAN DYNO TEST"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    ws1["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws1.merge_cells("A1:E1")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    # Header info kendaraan & sesi
    ws1["A3"] = "IDENTITAS KENDARAAN"
    ws1["A3"].font = Font(bold=True, color="000000")
    ws1["D3"] = "INFORMASI PENGUJIAN"
    ws1["D3"].font = Font(bold=True, color="000000")

    info_rows = [
        ("Nomor Uji (KIR)", vehicle.test_number if vehicle else "—", "ID Sesi", f"#{result.session_id}"),
        ("Nomor Rangka (VIN)", vehicle.vin if vehicle else "—", "Penguji", session.inspector_name if session else "—"),
        ("Nomor Polisi", vehicle.license_plate if vehicle else "—", "Waktu Uji", str(session.tested_at) if session and session.tested_at else datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Merk & Tipe", vehicle.brand_model if vehicle else "—", "Kategori", vehicle.vehicle_category if vehicle else "Roda 2"),
        ("Bobot Kendaraan", f"{vehicle.vehicle_weight_kg:.1f} kg" if vehicle else "150.0 kg", "Mode Uji", "DYNO TEST"),
    ]

    for idx, (lbl1, val1, lbl2, val2) in enumerate(info_rows, start=4):
        ws1[f"A{idx}"] = lbl1
        ws1[f"A{idx}"].font = Font(color="475569")
        ws1[f"B{idx}"] = val1
        ws1[f"B{idx}"].font = Font(bold=True)
        ws1[f"D{idx}"] = lbl2
        ws1[f"D{idx}"].font = Font(color="475569")
        ws1[f"E{idx}"] = val2
        ws1[f"E{idx}"].font = Font(bold=True)

    # Metric box table (PEAK MONITOR)
    start_r = 10
    ws1[f"A{start_r}"] = "PARAMETER"
    ws1[f"B{start_r}"] = "NILAI PUNCAK (PEAK)"
    ws1[f"C{start_r}"] = "SATUAN"
    ws1[f"D{start_r}"] = "RPM SAAT PEAK"
    for col_c in ["A", "B", "C", "D"]:
        cell = ws1[f"{col_c}{start_r}"]
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[start_r].height = 24

    peak_data = [
        ("Daya Mesin Maksimum (Power)", f"{result.max_power_hp:.2f}", "HP", f"{result.rpm_at_peak_power:.0f} RPM"),
        ("Torsi Maksimum (Torque)", f"{result.max_torque_nm:.1f}", "Nm", f"{result.rpm_at_peak_torque:.0f} RPM"),
        ("Kecepatan Puncak (Top Speed)", f"{result.max_speed_kmh:.1f}", "km/h", "—"),
        ("RPM Maksimum Mesin", f"{result.max_rpm:.0f}", "RPM", "—"),
    ]

    for r_idx, (p_lbl, p_val, p_unit, p_rpm) in enumerate(peak_data, start=start_r + 1):
        ws1[f"A{r_idx}"] = p_lbl
        ws1[f"B{r_idx}"] = p_val
        ws1[f"C{r_idx}"] = p_unit
        ws1[f"D{r_idx}"] = p_rpm
        for c_char in ["A", "B", "C", "D"]:
            cell = ws1[f"{c_char}{r_idx}"]
            cell.border = _thin_border()
            if c_char in ("B", "C", "D"):
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)

    _autofit_columns(ws1)

    # ── SHEET 2: TELEMETRI TIME SERIES ───────────────────────────────────────
    if result.raw_time_series:
        ws2 = wb.create_sheet(title="Telemetri Real-time")
        ws2.views.sheetView[0].showGridLines = True

        headers = ["Waktu (s)", "RPM", "Torsi (Nm)", "Daya Mesin (HP)", "Kecepatan (km/h)"]
        ws2.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 24

        for item in result.raw_time_series:
            ws2.append([
                item.get("t", 0.0),
                item.get("rpm", 0),
                item.get("torque", 0.0),
                item.get("power", 0.0),
                item.get("speed", 0.0),
            ])

        _autofit_columns(ws2)

    wb.save(filepath)
    return True


def export_brake_to_excel(
    filepath: str,
    session: Optional[TestSession],
    vehicle: Optional[Vehicle],
    result: BrakeResult,
) -> bool:
    """Export hasil pengujian Brake Test ke file Excel."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Ringkasan Brake Test"
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1["A1"] = "AUTO-TECH SYSTEMS — LAPORAN PENGUJIAN REM & LAMPU"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    ws1["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws1.merge_cells("A1:E1")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    # Identitas
    ws1["A3"] = "IDENTITAS KENDARAAN"
    ws1["A3"].font = Font(bold=True, color="000000")
    ws1["D3"] = "INFORMASI PENGUJIAN"
    ws1["D3"].font = Font(bold=True, color="000000")

    info_rows = [
        ("Nomor Uji (KIR)", vehicle.test_number if vehicle else "—", "ID Sesi", f"#{result.session_id}"),
        ("Nomor Rangka (VIN)", vehicle.vin if vehicle else "—", "Penguji", session.inspector_name if session else "—"),
        ("Nomor Polisi", vehicle.license_plate if vehicle else "—", "Waktu Uji", str(session.tested_at) if session and session.tested_at else datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Bobot Uji (kg)", f"{vehicle.vehicle_weight_kg:.1f} kg" if vehicle else "150.0 kg", "Mode Uji", "BRAKE TEST"),
    ]

    for idx, (lbl1, val1, lbl2, val2) in enumerate(info_rows, start=4):
        ws1[f"A{idx}"] = lbl1
        ws1[f"A{idx}"].font = Font(color="475569")
        ws1[f"B{idx}"] = val1
        ws1[f"B{idx}"].font = Font(bold=True)
        ws1[f"D{idx}"] = lbl2
        ws1[f"D{idx}"].font = Font(color="475569")
        ws1[f"E{idx}"] = val2
        ws1[f"E{idx}"].font = Font(bold=True)

    # Hasil Pengujian Table
    start_r = 9
    headers = ["PARAMETER PENGUJIAN", "HASIL UKUR", "STANDAR MINIMAL", "STATUS KELULUSAN"]
    for c_idx, h_text in enumerate(headers, start=1):
        cell = ws1.cell(row=start_r, column=c_idx, value=h_text)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[start_r].height = 24

    b_pass = result.brake_pass_status == EvaluationStatus.PASS
    l_pass = result.lux_pass_status == EvaluationStatus.PASS
    o_pass = result.overall_status == EvaluationStatus.PASS

    test_eval_rows = [
        ("Efisiensi Pengereman", f"{result.braking_efficiency_pct:.1f} %", "≥ 50.0 %", "LULUS" if b_pass else "TIDAK LULUS", b_pass),
        ("Gaya Pengereman Puncak", f"{result.peak_braking_force_n:,.0f} N", "—", "TERCATAT", True),
        ("Waktu Pengereman", f"{result.braking_time_s:.2f} s", "≤ 4.00 s", "LULUS" if result.braking_time_s <= 4.0 else "TIDAK LULUS", result.braking_time_s <= 4.0),
        ("Intensitas Cahaya Lampu", f"{result.lux_intensity:,.0f} Lux", "≥ 12,000 Lux", "LULUS" if l_pass else "TIDAK LULUS", l_pass),
        ("KESIMPULAN AKHIR", "LULUS KESELURUHAN" if o_pass else "TIDAK LULUS", "Semua Uji Lulus", "PASS" if o_pass else "FAIL", o_pass),
    ]

    for r_offset, (p_name, p_val, p_std, p_stat, is_ok) in enumerate(test_eval_rows, start=start_r + 1):
        ws1[f"A{r_offset}"] = p_name
        ws1[f"B{r_offset}"] = p_val
        ws1[f"C{r_offset}"] = p_std
        ws1[f"D{r_offset}"] = p_stat

        for c_char in ["A", "B", "C", "D"]:
            cell = ws1[f"{c_char}{r_offset}"]
            cell.border = _thin_border()
            if c_char in ("B", "C", "D"):
                cell.alignment = Alignment(horizontal="center")

        stat_cell = ws1[f"D{r_offset}"]
        stat_cell.font = Font(bold=True, color=COLOR_PASS_TEXT if is_ok else COLOR_FAIL_TEXT)
        stat_cell.fill = PatternFill("solid", fgColor=COLOR_PASS_BG if is_ok else COLOR_FAIL_BG)

    _autofit_columns(ws1)

    # ── SHEET 2: TELEMETRI ──────────────────────────────────────────────────
    if result.raw_time_series:
        ws2 = wb.create_sheet(title="Telemetri Rem")
        ws2.views.sheetView[0].showGridLines = True
        h2 = ["Waktu (s)", "Kecepatan (km/h)", "Gaya Rem (N)", "Waktu Pengereman (s)", "Lux"]
        ws2.append(h2)
        for col_idx in range(1, len(h2) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 24

        for item in result.raw_time_series:
            ws2.append([
                item.get("t", 0.0),
                item.get("speed", 0.0),
                item.get("brake_force", item.get("braking_force_n", 0.0)),
                item.get("braking_time_s", 0.0),
                item.get("lux", 0.0),
            ])
        _autofit_columns(ws2)

    wb.save(filepath)
    return True


def export_history_to_excel(filepath: str, rows: List[Dict[str, Any]]) -> bool:
    """Export daftar riwayat pengujian ke file Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Pengujian"
    ws.views.sheetView[0].showGridLines = True

    # Title
    ws["A1"] = "AUTO-TECH SYSTEMS — REKAPITULASI RIWAYAT PENGUJIAN"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_TEXT)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    headers = ["Tanggal & Waktu", "Nomor Uji", "Nomor Rangka", "Nama Penguji", "Mode Uji", "Status"]
    ws.append([])
    ws.append(headers)
    header_row = 3
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 24

    for r in rows:
        mode_val = r.get("test_mode", "—")
        if hasattr(mode_val, "value"):
            mode_val = mode_val.value
        ws.append([
            str(r.get("tested_at", "—")),
            str(r.get("test_number", "—")),
            str(r.get("vin", "—")),
            str(r.get("inspector_name", "—")),
            str(mode_val),
            "Selesai",
        ])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = _thin_border()

    _autofit_columns(ws)
    wb.save(filepath)
    return True
